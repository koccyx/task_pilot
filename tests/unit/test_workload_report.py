"""Tests for workload report generation."""

from datetime import date, timedelta
from pathlib import Path

import pytest
from openpyxl import load_workbook

from chat_bot.models import RouteResult
from chat_bot.reports.workload_report import WorkloadReportService


class FakeKaitenClient:
    """Minimal async Kaiten client for report tests."""

    def __init__(self) -> None:
        self.soon = (date.today() + timedelta(days=2)).isoformat()
        self.overdue = (date.today() - timedelta(days=1)).isoformat()
        self.requested_endpoints: list[str] = []

    async def get(self, endpoint: str) -> dict:
        self.requested_endpoints.append(endpoint)

        if endpoint == "spaces":
            return {"spaces": [{"id": 100, "title": "jmlc"}]}
        if endpoint == "spaces/100/boards":
            return {"boards": [{"id": 200, "title": "основная доска"}]}
        if endpoint.startswith("spaces/100/users?"):
            return {
                "users": [
                    {
                        "id": 1,
                        "full_name": "Анна Иванова",
                        "email": "anna@example.test",
                    },
                    {
                        "id": 2,
                        "full_name": "Иван Петров",
                        "email": "ivan@example.test",
                    },
                    {
                        "id": 3,
                        "full_name": "Мария Сидорова",
                        "email": "maria@example.test",
                    },
                ]
            }
        if endpoint.startswith("cards?"):
            return {
                "cards": [
                    {
                        "id": 101,
                        "title": "Сделать API",
                        "owner_id": 1,
                        "due_date": self.soon,
                        "asap": True,
                        "board": {"title": "основная доска"},
                        "column": {"name": "In Progress"},
                    },
                    {
                        "id": 102,
                        "title": "Починить баг",
                        "owner": {"id": 2, "full_name": "Иван Петров"},
                        "due_date": self.overdue,
                        "asap": False,
                    },
                    {
                        "id": 103,
                        "title": "Разобрать входящие",
                    },
                ]
            }
        raise AssertionError(f"Unexpected endpoint: {endpoint}")


def test_detects_report_requests() -> None:
    assert WorkloadReportService.is_workload_report_request("дай отчет")
    assert WorkloadReportService.is_workload_report_request(
        "кто свободен, кому можно дать задач?"
    )
    assert not WorkloadReportService.is_workload_report_request("покажи мои задачи")


@pytest.mark.asyncio
async def test_generate_creates_xlsx_with_summary() -> None:
    client = FakeKaitenClient()
    service = WorkloadReportService(client=client)

    result = await service.generate()

    assert isinstance(result, RouteResult)
    assert result.document_path is not None
    assert result.document_path.exists()
    assert result.document_path.suffix == ".xlsx"

    workbook = load_workbook(result.document_path)
    summary = workbook["Загруженность"]
    recommendations = workbook["Рекомендации"]
    problems = workbook["Проблемные карточки"]
    details = workbook["Карточки"]

    assert summary["A1"].value == "Отчет по загруженности команды"
    assert summary["A3"].value == "Пространство: jmlc; доска: основная доска"
    assignees = {summary.cell(row=row, column=1).value for row in range(6, 10)}
    assert {
        "Анна Иванова",
        "Иван Петров",
        "Мария Сидорова",
        "Без ответственного",
    } <= assignees
    assert summary["I5"].value == "Риск"

    recommendation_values = {
        recommendations.cell(row=row, column=2).value
        for row in range(6, recommendations.max_row + 1)
    }
    assert "Мария Сидорова" in recommendation_values
    assert problems["A1"].value == "Проблемные карточки"
    problem_reasons = {
        problems.cell(row=row, column=1).value for row in range(6, problems.max_row + 1)
    }
    assert any("просрочена" in str(reason) for reason in problem_reasons)
    assert any("без ответственного" in str(reason) for reason in problem_reasons)
    assert details.max_row == 4
    assert details["I1"].value == "Без дедлайна"
    assert "spaces/100/boards" in client.requested_endpoints
    assert any(
        "space_id=100" in endpoint and "board_id=200" in endpoint
        for endpoint in client.requested_endpoints
    )

    Path(result.document_path).unlink(missing_ok=True)
