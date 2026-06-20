"""Excel workload report generation for Kaiten cards."""

from __future__ import annotations

import re
import tempfile
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from chat_bot.logging_config import get_logger
from chat_bot.mcp_server.client.kaiten_client import KaitenClient, get_kaiten_client
from chat_bot.models import RouteResult

logger = get_logger(__name__)


@dataclass(frozen=True)
class Assignee:
    """Normalized card assignee."""

    user_id: Optional[int]
    name: str
    email: str = ""


@dataclass(frozen=True)
class WorkloadRow:
    """Aggregated workload for one assignee."""

    assignee: Assignee
    active_cards: int
    urgent_cards: int
    overdue_cards: int
    due_soon_cards: int
    no_due_date_cards: int
    workload_score: float
    risk_level: str
    recommendation: str


class WorkloadReportService:
    """Build an Excel report with team workload by Kaiten assignee."""

    def __init__(self, client: Optional[KaitenClient] = None) -> None:
        self._client = client

    @property
    def client(self) -> KaitenClient:
        """Lazily initialize Kaiten client so router construction is side-effect free."""
        if self._client is None:
            self._client = get_kaiten_client()
        return self._client

    @staticmethod
    def is_workload_report_request(text: str) -> bool:
        """Return true when text asks for a workload Excel report."""
        normalized = text.lower().replace("ё", "е")
        if re.search(r"\b(отчет|report)\b", normalized):
            return True
        workload_terms = (
            "загружен",
            "занят",
            "свобод",
            "кому можно",
            "на кого можно",
            "побольше задач",
        )
        task_terms = ("задач", "карточ", "работ")
        return any(term in normalized for term in workload_terms) and any(
            term in normalized for term in task_terms
        )

    async def generate(self) -> RouteResult:
        """Fetch Kaiten data and return a ready-to-send Excel report."""
        users = await self._fetch_all("users")
        cards = await self._fetch_all("cards", {"condition": 1})

        user_by_id = {
            user.get("id"): user
            for user in users
            if isinstance(user, dict) and user.get("id") is not None
        }
        rows, details, problems = self._build_rows(cards, user_by_id)

        output_path = self._write_workbook(rows, details, problems)
        text = (
            "Готово. Собрал Excel-отчет по загруженности команды: "
            f"{len(rows)} исполнителей, {len(details)} активных назначений, "
            f"{len(problems)} проблемных карточек."
        )
        return RouteResult(
            text=text,
            document_path=output_path,
            document_filename=output_path.name,
        )

    async def _fetch_all(
        self,
        endpoint: str,
        params: Optional[dict[str, Any]] = None,
        page_size: int = 100,
        max_pages: int = 20,
    ) -> list[dict[str, Any]]:
        """Fetch paginated Kaiten collections that accept limit/offset."""
        items: list[dict[str, Any]] = []
        base_params = dict(params or {})

        for page in range(max_pages):
            query = {**base_params, "limit": page_size, "offset": page * page_size}
            response = await self.client.get(self._endpoint_with_query(endpoint, query))
            page_items = self._extract_collection(response, endpoint)
            items.extend(page_items)
            if len(page_items) < page_size:
                break

        logger.info("Fetched %d Kaiten %s for workload report", len(items), endpoint)
        return items

    @staticmethod
    def _endpoint_with_query(endpoint: str, params: dict[str, Any]) -> str:
        query = "&".join(f"{key}={value}" for key, value in params.items())
        return f"{endpoint}?{query}" if query else endpoint

    @staticmethod
    def _extract_collection(response: Any, endpoint: str) -> list[dict[str, Any]]:
        if isinstance(response, list):
            return [item for item in response if isinstance(item, dict)]
        if isinstance(response, dict):
            for key in (endpoint, "data", "cards", "users"):
                value = response.get(key)
                if isinstance(value, list):
                    return [item for item in value if isinstance(item, dict)]
            return [response] if response else []
        return []

    def _build_rows(
        self,
        cards: Iterable[dict[str, Any]],
        user_by_id: dict[Any, dict[str, Any]],
    ) -> tuple[list[WorkloadRow], list[dict[str, Any]], list[dict[str, Any]]]:
        today = date.today()
        aggregates: dict[str, dict[str, Any]] = {}
        details: list[dict[str, Any]] = []
        problems: list[dict[str, Any]] = []

        for card in cards:
            assignees = self._extract_assignees(card, user_by_id)
            due_date = self._parse_date(card.get("due_date"))
            is_overdue = bool(due_date and due_date < today)
            is_due_soon = bool(due_date and 0 <= (due_date - today).days <= 7)
            has_no_due_date = due_date is None
            is_urgent = bool(card.get("asap"))
            has_no_assignee = all(
                assignee.name == "Без ответственного" for assignee in assignees
            )

            problem_reason = self._problem_reason(
                is_overdue=is_overdue,
                is_urgent=is_urgent,
                has_no_assignee=has_no_assignee,
                has_no_due_date=has_no_due_date,
            )
            if problem_reason:
                problems.append(
                    {
                        "reason": problem_reason,
                        "assignee": ", ".join(assignee.name for assignee in assignees),
                        "card_id": card.get("id", ""),
                        "title": card.get("title", ""),
                        "board": self._name_from_nested(card.get("board"))
                        or card.get("board_id", ""),
                        "column": self._name_from_nested(card.get("column"))
                        or card.get("column_id", ""),
                        "due_date": due_date.isoformat() if due_date else "",
                    }
                )

            for assignee in assignees:
                key = (
                    str(assignee.user_id)
                    if assignee.user_id is not None
                    else assignee.name
                )
                stats = aggregates.setdefault(
                    key,
                    {
                        "assignee": assignee,
                        "active": 0,
                        "urgent": 0,
                        "overdue": 0,
                        "due_soon": 0,
                        "no_due_date": 0,
                    },
                )
                stats["active"] += 1
                stats["urgent"] += int(is_urgent)
                stats["overdue"] += int(is_overdue)
                stats["due_soon"] += int(is_due_soon)
                stats["no_due_date"] += int(has_no_due_date)
                details.append(
                    {
                        "assignee": assignee.name,
                        "card_id": card.get("id", ""),
                        "title": card.get("title", ""),
                        "board": self._name_from_nested(card.get("board"))
                        or card.get("board_id", ""),
                        "column": self._name_from_nested(card.get("column"))
                        or card.get("column_id", ""),
                        "due_date": due_date.isoformat() if due_date else "",
                        "urgent": "да" if is_urgent else "нет",
                        "overdue": "да" if is_overdue else "нет",
                        "no_due_date": "да" if has_no_due_date else "нет",
                    }
                )

        for user in user_by_id.values():
            assignee = self._assignee_from_user(user)
            key = (
                str(assignee.user_id) if assignee.user_id is not None else assignee.name
            )
            aggregates.setdefault(
                key,
                {
                    "assignee": assignee,
                    "active": 0,
                    "urgent": 0,
                    "overdue": 0,
                    "due_soon": 0,
                    "no_due_date": 0,
                },
            )

        rows: list[WorkloadRow] = []
        for stats in aggregates.values():
            score = (
                float(stats["active"])
                + float(stats["urgent"]) * 0.5
                + float(stats["overdue"]) * 0.75
                + float(stats["due_soon"]) * 0.25
            )
            rows.append(
                WorkloadRow(
                    assignee=stats["assignee"],
                    active_cards=stats["active"],
                    urgent_cards=stats["urgent"],
                    overdue_cards=stats["overdue"],
                    due_soon_cards=stats["due_soon"],
                    no_due_date_cards=stats["no_due_date"],
                    workload_score=round(score, 2),
                    risk_level=self._risk_level(
                        score=score,
                        overdue_cards=stats["overdue"],
                        urgent_cards=stats["urgent"],
                    ),
                    recommendation=self._recommendation(
                        score=score,
                        overdue_cards=stats["overdue"],
                        active_cards=stats["active"],
                    ),
                )
            )

        rows.sort(
            key=lambda row: (row.workload_score, row.active_cards, row.assignee.name)
        )
        details.sort(key=lambda item: (item["assignee"], str(item["card_id"])))
        problems.sort(key=lambda item: (item["reason"], str(item["card_id"])))
        return rows, details, problems

    @staticmethod
    def _recommendation(score: float, overdue_cards: int, active_cards: int) -> str:
        if active_cards == 0:
            return "Свободен для новых задач"
        if score <= 2:
            return "Можно добавить задачи"
        if overdue_cards > 0 and score > 5:
            return "Не добавлять задачи, сначала снять просрочки"
        if score <= 5:
            return "Нормальная загрузка"
        return "Высокая загрузка"

    @staticmethod
    def _risk_level(score: float, overdue_cards: int, urgent_cards: int) -> str:
        if overdue_cards >= 2 or score > 7:
            return "Высокий"
        if overdue_cards or urgent_cards >= 2 or score > 4:
            return "Средний"
        return "Низкий"

    @staticmethod
    def _problem_reason(
        is_overdue: bool,
        is_urgent: bool,
        has_no_assignee: bool,
        has_no_due_date: bool,
    ) -> str:
        reasons = []
        if is_overdue:
            reasons.append("просрочена")
        if is_urgent:
            reasons.append("срочная")
        if has_no_assignee:
            reasons.append("без ответственного")
        if has_no_due_date:
            reasons.append("без дедлайна")
        return ", ".join(reasons)

    def _extract_assignees(
        self,
        card: dict[str, Any],
        user_by_id: dict[Any, dict[str, Any]],
    ) -> list[Assignee]:
        assignees: list[Assignee] = []

        for key in ("owner", "assignee", "responsible", "responsible_user"):
            assignees.extend(self._assignees_from_value(card.get(key), user_by_id))

        for key in ("owners", "assignees", "members", "responsibles"):
            value = card.get(key)
            if isinstance(value, list):
                for item in value:
                    assignees.extend(self._assignees_from_value(item, user_by_id))

        for key in ("owner_id", "assignee_id", "responsible_id"):
            value = card.get(key)
            if value is not None:
                assignees.extend(self._assignees_from_value(value, user_by_id))

        unique: dict[str, Assignee] = {}
        for assignee in assignees:
            key = (
                str(assignee.user_id) if assignee.user_id is not None else assignee.name
            )
            unique[key] = assignee

        if not unique:
            return [Assignee(user_id=None, name="Без ответственного")]
        return list(unique.values())

    def _assignees_from_value(
        self,
        value: Any,
        user_by_id: dict[Any, dict[str, Any]],
    ) -> list[Assignee]:
        if value is None:
            return []
        if isinstance(value, dict):
            return [self._assignee_from_user(value)]
        if isinstance(value, int):
            user = user_by_id.get(value)
            if user:
                return [self._assignee_from_user(user)]
            return [Assignee(user_id=value, name=f"User {value}")]
        if isinstance(value, str) and value.strip():
            return [Assignee(user_id=None, name=value.strip())]
        return []

    @staticmethod
    def _assignee_from_user(user: dict[str, Any]) -> Assignee:
        user_id = user.get("id")
        name = (
            user.get("full_name")
            or user.get("name")
            or user.get("username")
            or user.get("email")
            or f"User {user_id}"
        )
        return Assignee(
            user_id=user_id if isinstance(user_id, int) else None,
            name=str(name),
            email=str(user.get("email") or ""),
        )

    @staticmethod
    def _name_from_nested(value: Any) -> str:
        if isinstance(value, dict):
            return str(value.get("title") or value.get("name") or "")
        return ""

    @staticmethod
    def _parse_date(value: Any) -> Optional[date]:
        if not isinstance(value, str) or not value:
            return None
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}", value):
            return datetime.strptime(value, "%Y-%m-%d").date()
        try:
            normalized = value.replace("Z", "+00:00")
            return datetime.fromisoformat(normalized).date()
        except ValueError:
            return None

    def _write_workbook(
        self,
        rows: list[WorkloadRow],
        details: list[dict[str, Any]],
        problems: list[dict[str, Any]],
    ) -> Path:
        workbook = Workbook()
        summary = workbook.active
        summary.title = "Загруженность"
        recommendations_sheet = workbook.create_sheet("Рекомендации")
        problems_sheet = workbook.create_sheet("Проблемные карточки")
        details_sheet = workbook.create_sheet("Карточки")

        self._fill_summary(summary, rows)
        self._fill_recommendations(recommendations_sheet, rows)
        self._fill_problems(problems_sheet, problems)
        self._fill_details(details_sheet, details)

        output_dir = Path(tempfile.gettempdir()) / "kaiten_workload_reports"
        output_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = output_dir / f"kaiten_workload_report_{timestamp}.xlsx"
        workbook.save(output_path)
        return output_path

    def _fill_summary(self, sheet: Any, rows: list[WorkloadRow]) -> None:
        headers = [
            "Исполнитель",
            "Email",
            "Активные задачи",
            "Срочные",
            "Просроченные",
            "Срок до 7 дней",
            "Без дедлайна",
            "Индекс загрузки",
            "Риск",
            "Рекомендация",
        ]
        sheet.append(["Отчет по загруженности команды"])
        sheet.append([f"Сформирован: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
        sheet.append([])
        sheet.append(headers)

        for row in rows:
            sheet.append(
                [
                    row.assignee.name,
                    row.assignee.email,
                    row.active_cards,
                    row.urgent_cards,
                    row.overdue_cards,
                    row.due_soon_cards,
                    row.no_due_date_cards,
                    row.workload_score,
                    row.risk_level,
                    row.recommendation,
                ]
            )

        self._style_sheet(sheet, header_row=4)
        sheet.freeze_panes = "A5"

    def _fill_recommendations(self, sheet: Any, rows: list[WorkloadRow]) -> None:
        sheet.append(["Рекомендации по распределению задач"])
        sheet.append([f"Сформирован: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
        sheet.append([])
        sheet.append(["Категория", "Исполнитель", "Индекс", "Риск", "Комментарий"])

        free_rows = [row for row in rows if row.active_cards == 0]
        candidates = [
            row
            for row in rows
            if row.active_cards > 0
            and row.workload_score <= 2
            and not row.overdue_cards
        ]
        overloaded = [
            row
            for row in rows
            if row.risk_level == "Высокий"
            or row.recommendation.startswith("Не добавлять")
        ]

        for row in free_rows[:10]:
            sheet.append(
                [
                    "Кандидат на новые задачи",
                    row.assignee.name,
                    row.workload_score,
                    row.risk_level,
                    "Нет активных задач",
                ]
            )
        for row in candidates[:10]:
            sheet.append(
                [
                    "Можно добавить задачи",
                    row.assignee.name,
                    row.workload_score,
                    row.risk_level,
                    f"Активных задач: {row.active_cards}",
                ]
            )
        for row in overloaded[:10]:
            sheet.append(
                [
                    "Не нагружать",
                    row.assignee.name,
                    row.workload_score,
                    row.risk_level,
                    (
                        f"Просрочено: {row.overdue_cards}; "
                        f"срочных: {row.urgent_cards}"
                    ),
                ]
            )

        if sheet.max_row == 4:
            sheet.append(["Нет рекомендаций", "", "", "", "Данных недостаточно"])

        self._style_sheet(sheet, header_row=4)
        sheet.freeze_panes = "A5"

    def _fill_problems(self, sheet: Any, problems: list[dict[str, Any]]) -> None:
        headers = [
            "Причина",
            "Исполнитель",
            "ID карточки",
            "Название",
            "Доска",
            "Колонка",
            "Дедлайн",
        ]
        sheet.append(["Проблемные карточки"])
        sheet.append([f"Сформирован: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
        sheet.append([])
        sheet.append(headers)

        for item in problems:
            sheet.append(
                [
                    item["reason"],
                    item["assignee"],
                    item["card_id"],
                    item["title"],
                    item["board"],
                    item["column"],
                    item["due_date"],
                ]
            )

        if not problems:
            sheet.append(["Нет проблемных карточек", "", "", "", "", "", ""])

        self._style_sheet(sheet, header_row=4)
        sheet.freeze_panes = "A5"

    def _fill_details(self, sheet: Any, details: list[dict[str, Any]]) -> None:
        headers = [
            "Исполнитель",
            "ID карточки",
            "Название",
            "Доска",
            "Колонка",
            "Дедлайн",
            "Срочная",
            "Просрочена",
            "Без дедлайна",
        ]
        sheet.append(headers)
        for item in details:
            sheet.append(
                [
                    item["assignee"],
                    item["card_id"],
                    item["title"],
                    item["board"],
                    item["column"],
                    item["due_date"],
                    item["urgent"],
                    item["overdue"],
                    item["no_due_date"],
                ]
            )
        self._style_sheet(sheet, header_row=1)
        sheet.freeze_panes = "A2"

    @staticmethod
    def _style_sheet(sheet: Any, header_row: int) -> None:
        title_fill = PatternFill("solid", fgColor="1F4E78")
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        title_font = Font(color="FFFFFF", bold=True, size=14)
        header_font = Font(bold=True)

        if header_row > 1:
            sheet["A1"].font = title_font
            sheet["A1"].fill = title_fill
            sheet.merge_cells(
                start_row=1,
                start_column=1,
                end_row=1,
                end_column=sheet.max_column,
            )

        for cell in sheet[header_row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for column_cells in sheet.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            width = min(max(length + 2, 12), 45)
            letter = get_column_letter(column_cells[0].column)
            sheet.column_dimensions[letter].width = width

        sheet.auto_filter.ref = sheet.dimensions
