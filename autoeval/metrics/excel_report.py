"""Build an Excel metrics report from an autoeval eval_result JSON file."""

from __future__ import annotations

import argparse
import json
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_EVAL_RESULT = (
    PROJECT_ROOT / "autoeval" / "eval" / "results" / "eval_result_agent_latest.json"
)
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "autoeval" / "metrics" / "results"


class MetricsError(RuntimeError):
    """Raised when the metrics report cannot be generated."""


@dataclass(frozen=True)
class GroupMetrics:
    """Aggregated metrics for one criteria group type."""

    criteria_type: str
    passed_criteria: int
    total_criteria: int
    passed_groups: int
    total_groups: int

    @property
    def micro(self) -> float | None:
        """Successful criteria divided by all criteria."""
        if self.total_criteria == 0:
            return None
        return self.passed_criteria / self.total_criteria

    @property
    def macro(self) -> float | None:
        """Fully successful groups divided by all groups."""
        if self.total_groups == 0:
            return None
        return self.passed_groups / self.total_groups


def utc_now() -> str:
    """Return an ISO-8601 UTC timestamp with a Z suffix."""
    return (
        datetime.now(timezone.utc)
        .replace(microsecond=0)
        .isoformat()
        .replace("+00:00", "Z")
    )


def load_json(path: Path) -> dict[str, Any]:
    """Load a JSON object."""
    with path.open("r", encoding="utf-8") as file:
        data = json.load(file)
    if not isinstance(data, dict):
        raise MetricsError(f"Expected JSON object in {path}")
    return data


def output_path(output_dir: Path, eval_result_path: Path) -> Path:
    """Build a default output path for the Excel report."""
    stem = eval_result_path.stem.replace("eval_result", "metrics")
    if stem == eval_result_path.stem:
        stem = f"metrics_{utc_now().replace(':', '-')}"
    return output_dir / f"{stem}.xlsx"


def eval_entries(eval_result: dict[str, Any]) -> list[dict[str, Any]]:
    """Return normalized eval result entries."""
    entries = eval_result.get("results")
    if not isinstance(entries, list):
        raise MetricsError('Eval result must contain a "results" array')
    return [entry for entry in entries if isinstance(entry, dict)]


def criterion_results(entry: dict[str, Any]) -> list[dict[str, Any]]:
    """Return normalized criterion results for one criteria group."""
    results = entry.get("results")
    if not isinstance(results, list):
        return []
    return [result for result in results if isinstance(result, dict)]


def group_is_success(results: list[dict[str, Any]]) -> bool:
    """Return whether every criterion in a group passed."""
    return bool(results) and all(bool(result.get("is_success")) for result in results)


def aggregate_metrics(entries: list[dict[str, Any]]) -> list[GroupMetrics]:
    """Aggregate micro and macro inputs by criteria_type and overall."""
    by_type: dict[str, dict[str, int]] = {}
    overall = {
        "passed_criteria": 0,
        "total_criteria": 0,
        "passed_groups": 0,
        "total_groups": 0,
    }

    for entry in entries:
        criteria_type = str(entry.get("criteria_type") or "unknown")
        results = criterion_results(entry)
        passed_criteria = sum(1 for result in results if bool(result.get("is_success")))
        total_criteria = len(results)
        passed_groups = 1 if group_is_success(results) else 0

        bucket = by_type.setdefault(
            criteria_type,
            {
                "passed_criteria": 0,
                "total_criteria": 0,
                "passed_groups": 0,
                "total_groups": 0,
            },
        )
        bucket["passed_criteria"] += passed_criteria
        bucket["total_criteria"] += total_criteria
        bucket["passed_groups"] += passed_groups
        bucket["total_groups"] += 1

        overall["passed_criteria"] += passed_criteria
        overall["total_criteria"] += total_criteria
        overall["passed_groups"] += passed_groups
        overall["total_groups"] += 1

    metrics = [
        GroupMetrics(criteria_type=criteria_type, **values)
        for criteria_type, values in sorted(by_type.items())
    ]
    metrics.append(GroupMetrics(criteria_type="overall", **overall))
    return metrics


def percent(value: float | None) -> float | None:
    """Convert a ratio to a percentage value."""
    if value is None:
        return None
    return round(value * 100, 2)


def append_header(sheet: Any, headers: list[str]) -> None:
    """Append and style a header row."""
    sheet.append(headers)
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")


def autosize(sheet: Any) -> None:
    """Adjust column widths based on cell contents."""
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)


def style_score_columns(sheet: Any, columns: tuple[int, ...]) -> None:
    """Format percentage score columns."""
    for row in sheet.iter_rows(min_row=2):
        for column_index in columns:
            cell = row[column_index - 1]
            if isinstance(cell.value, (int, float)):
                cell.number_format = '0.00"%"'


def write_summary_sheet(workbook: Workbook, eval_result: dict[str, Any]) -> None:
    """Write summary metrics sheet."""
    sheet = workbook.active
    sheet.title = "summary"
    append_header(
        sheet,
        [
            "metric",
            "criteria_type",
            "passed",
            "total",
            "score_percent",
        ],
    )

    for metric in aggregate_metrics(eval_entries(eval_result)):
        sheet.append(
            [
                "micro" if metric.criteria_type != "overall" else "overall_micro",
                metric.criteria_type,
                metric.passed_criteria,
                metric.total_criteria,
                percent(metric.micro),
            ]
        )
        sheet.append(
            [
                "macro" if metric.criteria_type != "overall" else "overall_macro",
                metric.criteria_type,
                metric.passed_groups,
                metric.total_groups,
                percent(metric.macro),
            ]
        )

    style_score_columns(sheet, (5,))
    autosize(sheet)


def write_by_case_sheet(workbook: Workbook, entries: list[dict[str, Any]]) -> None:
    """Write per-case group metrics."""
    sheet = workbook.create_sheet("by_case")
    append_header(
        sheet,
        [
            "case_id",
            "case_name",
            "criteria_type",
            "passed_criteria",
            "total_criteria",
            "micro_percent",
            "group_success",
        ],
    )

    for entry in entries:
        results = criterion_results(entry)
        total = len(results)
        passed = sum(1 for result in results if bool(result.get("is_success")))
        sheet.append(
            [
                entry.get("case_id"),
                entry.get("case_name"),
                entry.get("criteria_type"),
                passed,
                total,
                percent(passed / total) if total else None,
                group_is_success(results),
            ]
        )

    style_score_columns(sheet, (6,))
    autosize(sheet)


def write_failures_sheet(workbook: Workbook, entries: list[dict[str, Any]]) -> None:
    """Write failed criterion details."""
    sheet = workbook.create_sheet("failures")
    append_header(
        sheet,
        [
            "case_id",
            "case_name",
            "criteria_type",
            "criterion",
            "reason",
        ],
    )

    for entry in entries:
        for result in criterion_results(entry):
            if bool(result.get("is_success")):
                continue
            sheet.append(
                [
                    entry.get("case_id"),
                    entry.get("case_name"),
                    entry.get("criteria_type"),
                    result.get("criterion"),
                    result.get("reason"),
                ]
            )

    autosize(sheet)


def write_metadata_sheet(
    workbook: Workbook,
    eval_result: dict[str, Any],
    eval_result_path: Path,
) -> None:
    """Write report metadata."""
    sheet = workbook.create_sheet("metadata")
    append_header(sheet, ["field", "value"])
    for key in (
        "benchmark_id",
        "run_id",
        "eval_id",
        "eval_time",
        "source_run_result",
    ):
        sheet.append([key, eval_result.get(key)])
    sheet.append(["eval_result_path", str(eval_result_path)])
    sheet.append(["generated_at", utc_now()])
    sheet.append(["summary_json", json.dumps(eval_result.get("summary", {}), ensure_ascii=False)])
    sheet.append(["judge_json", json.dumps(eval_result.get("judge", {}), ensure_ascii=False)])
    autosize(sheet)


def build_workbook(eval_result: dict[str, Any], eval_result_path: Path) -> Workbook:
    """Build an Excel workbook for an eval result."""
    entries = eval_entries(eval_result)
    workbook = Workbook()
    write_summary_sheet(workbook, eval_result)
    write_by_case_sheet(workbook, entries)
    write_failures_sheet(workbook, entries)
    write_metadata_sheet(workbook, eval_result, eval_result_path)
    return workbook


def save_report(eval_result_path: Path, output: Path) -> Path:
    """Create and save an Excel metrics report."""
    eval_result = load_json(eval_result_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = build_workbook(eval_result, eval_result_path)
    workbook.save(output)
    return output


def parse_args() -> argparse.Namespace:
    """Parse CLI arguments."""
    parser = argparse.ArgumentParser(
        description="Create an Excel metrics report from autoeval eval_result JSON."
    )
    parser.add_argument(
        "--eval-result",
        type=Path,
        default=DEFAULT_EVAL_RESULT,
        help=f"Path to eval_result JSON. Default: {DEFAULT_EVAL_RESULT}",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Exact output .xlsx path. Defaults to autoeval/metrics/results/.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help=f"Directory for generated .xlsx files. Default: {DEFAULT_OUTPUT_DIR}",
    )
    return parser.parse_args()


def main() -> None:
    """CLI entrypoint."""
    args = parse_args()
    eval_result_path = args.eval_result.resolve()
    report_path = (
        args.output.resolve()
        if args.output
        else output_path(args.output_dir.resolve(), eval_result_path)
    )
    saved_path = save_report(eval_result_path, report_path)
    print(str(saved_path))


if __name__ == "__main__":
    main()
